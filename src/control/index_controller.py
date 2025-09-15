# moved from src/index_controller.py
from flask import Blueprint, render_template, request, flash, session, send_file, jsonify
from auth import login_required
from services import (
    load_skills, extract_text, extract_entities_ner, extract_basic_entities,
    extract_skills_spacy, calculate_semantic_similarity, save_cv_to_session,
    set_ranking_results, get_ranking_results, save_jd_to_session,
    get_jd_from_session, clear_jd_from_session, clear_user_uploads_directory,
    clear_session_data, get_user_uploads_dir, save_file_to_user_dir,
    get_user_file_path, list_user_files, cleanup_old_user_sessions
)
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from werkzeug.datastructures import FileStorage
from datetime import datetime
import os
import re
import json
import pandas as pd
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch


bp = Blueprint('index_bp', __name__)


def process_uploaded_resumes(resume_files):
    """
    Process uploaded CV files
    Return number of successfully uploaded files
    """
    uploaded_count = 0
    
    if not resume_files:
        return uploaded_count

    # Create uploads directory for user
    user_dir = get_user_uploads_dir()

    for resume_file in resume_files:
        if resume_file.filename.lower().endswith(('.pdf', '.docx')):
            try:
                # Save file to user directory
                resume_path = save_file_to_user_dir(resume_file, resume_file.filename)
                if resume_path:
                    # Extract information from CV
                    cv_data = extract_cv_data(resume_file.filename)
                    if cv_data:
                        # Save to session
                        save_cv_to_session(cv_data)
                        uploaded_count += 1
                        print(f"✅ Processed CV: {resume_file.filename}")
                    else:
                        print(f"❌ Failed to extract CV data: {resume_file.filename}")
            except Exception as e:
                print(f"Error processing file {resume_file.filename}: {e}")
                continue

    return uploaded_count


def extract_cv_data(filename):
    """Extract information from CV file"""
    try:
        # Use user-specific file path
        file_path = get_user_file_path(filename)
        
        # Extract text from file
        text = extract_text(file_path)
        if not text:
            return None

        # Extract entities
        entities = extract_entities_ner(text)
        emails, phones, years_exp = extract_basic_entities(text)
        skills = extract_skills_spacy(text)

        # Create CV data
        cv_data = {
            'filename': filename,
            'file_path': file_path,  # User-specific path
            'name': entities['PERSON'][0] if entities['PERSON'] else filename,
            'email': emails[0] if emails else "N/A",
            'phone': phones[0] if phones else "N/A", 
            'years_exp': years_exp,
            'skills': skills,
            'skills_text': ", ".join(skills[:10]),
            'text': text,  # Save full text for search
            'entities': entities,
            'emails': emails,
            'phones': phones
        }

        return cv_data
        
    except Exception as e:
        print(f"Error extracting CV data from {filename}: {e}")
        return None


def load_uploaded_cvs_from_session():
    """
    Load uploaded CVs from session
    Returns list of CV data
    """
    uploaded_cvs = []
    
    if 'uploaded_cvs' not in session or not session['uploaded_cvs']:
        return uploaded_cvs
    
    for cv in session['uploaded_cvs']:
        json_file_path = cv.get('json_file_path')
        if json_file_path and os.path.exists(json_file_path):
            try:
                with open(json_file_path, "r", encoding="utf-8") as json_file:
                    uploaded_cvs.append(json.load(json_file))
            except Exception as e:
                print(f"Error reading JSON file {json_file_path}: {e}")
                continue
    
    return uploaded_cvs


def prepare_resume_texts(uploaded_cvs):
    """
    Prepare text from CVs for analysis
    Returns list of CV texts
    """
    resume_texts = []
    
    for cv in uploaded_cvs:
        text_for_model = cv.get('text', '')
        
        # If no text, try reading from file again
        if not text_for_model:
            file_path = os.path.join('uploads', cv['filename'])
            if os.path.exists(file_path):
                try:
                    text_for_model = extract_text(file_path)
                except Exception as e:
                    print(f"Error reading file again {cv['filename']}: {e}")
        
        # Fallback: use skills_text
        if not text_for_model:
            text_for_model = cv.get('skills_text', '')
        
        resume_texts.append(text_for_model)
    
    return resume_texts


def calculate_tfidf_similarities(resume_texts, job_description):
    """
    Calculate TF-IDF similarity between JD and CVs
    Returns list of TF-IDF scores
    """
    try:
        tfidf_vectorizer = TfidfVectorizer(stop_words='english', max_features=10000)
        job_desc_vector = tfidf_vectorizer.fit_transform([job_description])
        
        tfidf_similarities = []
        for resume_text in resume_texts:
            resume_vector = tfidf_vectorizer.transform([resume_text])
            similarity = cosine_similarity(job_desc_vector, resume_vector)[0][0]
            tfidf_similarities.append(similarity)
        
        return tfidf_similarities
        
    except Exception as e:
        print(f"Error calculating TF-IDF: {e}")
        return [0.0] * len(resume_texts)


def calculate_semantic_similarities(resume_texts, job_description):
    """
    Calculate semantic similarity between JD and CVs
    Returns list of semantic scores
    """
    try:
        return calculate_semantic_similarity(resume_texts, job_description)
    except Exception as e:
        print(f"Error calculating semantic similarity: {e}")
        return [0.0] * len(resume_texts)


def calculate_skill_scores(uploaded_cvs, job_description):
    """
    Calculate skill scores between JD and CVs
    Returns list of skill scores
    """
    try:
        # Extract skills from JD
        jd_skills = set([s.strip().lower() for s in extract_skills_spacy(job_description)])
        
        # Extract skills from CVs
        cv_skills_sets = []
        for cv in uploaded_cvs:
            cv_skills = set([s.strip().lower() for s in cv.get('skills_text', '').split(',') if s.strip()])
            cv_skills_sets.append(cv_skills)
        
        # Calculate overlap
        overlaps = [len(cv_skills & jd_skills) for cv_skills in cv_skills_sets]
        max_overlap = max(overlaps) if overlaps else 1
        skill_scores = [(ov / max_overlap) * 100 if max_overlap > 0 else 0.0 for ov in overlaps]
        
        return skill_scores
        
    except Exception as e:
        print(f"Error calculating skill scores: {e}")
        return [0.0] * len(uploaded_cvs)


def create_ranked_resumes(uploaded_cvs, tfidf_similarities, semantic_similarities, skill_scores):
    """
    Create ranked CV list
    Returns list of CVs sorted by score
    """
    ranked_resumes = []
    
    for i, cv in enumerate(uploaded_cvs):
        # Calculate composite score (weights can be adjusted)
        combined_score = (0.4 * tfidf_similarities[i] +
                         0.5 * semantic_similarities[i] +
                         0.1 * (skill_scores[i] / 100.0)) * 100
        
        ranked_resumes.append({
            'rank': 0,  # Will be updated after sorting
            'filename': cv['filename'],
            'name': cv['name'],
            'email': cv['email'],
            'phone': cv['phone'],
            'years_exp': cv.get('years_exp', 0.0),
            'skills_text': cv['skills_text'],
            'tfidf_score': tfidf_similarities[i] * 100,
            'semantic_score': semantic_similarities[i] * 100,
            'skill_score': skill_scores[i],
            'combined_score': combined_score,
            'selected': False
        })
    
    # Sort by score in descending order
    ranked_resumes.sort(key=lambda x: x['combined_score'], reverse=True)
    
    # Update rank
    for i, resume in enumerate(ranked_resumes):
        resume['rank'] = i + 1
    
    return ranked_resumes


def analyze_resumes(job_description, uploaded_cvs):
    """
    Analyze and rank CVs based on Job Description
    Returns list of ranked CVs
    """
    try:
        # Prepare text from CVs
        resume_texts = prepare_resume_texts(uploaded_cvs)
        
        # Calculate different types of scores
        tfidf_similarities = calculate_tfidf_similarities(resume_texts, job_description)
        semantic_similarities = calculate_semantic_similarities(resume_texts, job_description)
        skill_scores = calculate_skill_scores(uploaded_cvs, job_description)
        
        # Create ranking list
        ranked_resumes = create_ranked_resumes(uploaded_cvs, tfidf_similarities, semantic_similarities, skill_scores)
        
        return ranked_resumes
        
    except Exception as e:
        print(f"Error analyzing CVs: {e}")
        return []


@bp.route('/', methods=['GET', 'POST'])
@login_required
def index():
    """
    Main route for index page
    Handle CV upload and analysis
    """
    load_skills()
    
    if request.method == 'POST':
        # Handle CV file uploads
        if 'resume_files' in request.files:
            resume_files = request.files.getlist('resume_files')
            uploaded_count = process_uploaded_resumes(resume_files)
            
            if uploaded_count > 0:
                flash(f'Successfully uploaded {uploaded_count} CV files.', 'success')
            else:
                flash('No CV files were uploaded successfully.', 'warning')
        
        # Get Job Description
        job_description = request.form.get('job_description', '').strip()
        
        if not job_description:
            flash('Please enter Job Description.', 'error')
            return render_template('index.html')
        
        # Load CVs from session
        uploaded_cvs = load_uploaded_cvs_from_session()
        
        if not uploaded_cvs:
            flash('No CVs have been uploaded yet. Please upload CVs first.', 'error')
            return render_template('index.html')
        
        # Analyze and rank CVs
        ranked_resumes = analyze_resumes(job_description, uploaded_cvs)
        
        if not ranked_resumes:
            flash('An error occurred while analyzing CVs.', 'error')
            return render_template('index.html')
        
        # Save results
        set_ranking_results(ranked_resumes)
        flash(f'Successfully analyzed and ranked {len(ranked_resumes)} CVs!', 'success')
    
    return render_template('index.html', results=get_ranking_results())


def get_cv_feedback_info(filename):
    """
    Get feedback information from CV's JSON file
    """
    try:
        # ✅ Use user-specific path instead of root uploads
        from services import get_user_uploads_dir
        user_dir = get_user_uploads_dir()
        cv_json_file = os.path.join(user_dir, f"{filename}.json")
        
        if not os.path.exists(cv_json_file):
            return {
                'total_feedback': 0,
                'latest_decision': None,
                'average_rating': None,
                'latest_feedback_time': None,
                'decision_breakdown': {}  # ✅ Add missing field
            }
        
        with open(cv_json_file, 'r', encoding='utf-8') as f:
            cv_data = json.load(f)
        
        feedback_history = cv_data.get('feedback_history', [])
        latest_feedback = cv_data.get('latest_feedback', None)
        
        # Calculate average rating
        ratings = [float(fb.get('human_rating', 0)) for fb in feedback_history 
                  if fb.get('human_rating') and str(fb.get('human_rating')).strip() != '']
        
        # Calculate decision breakdown
        decision_breakdown = {}
        for fb in feedback_history:
            decision = fb.get('human_decision', 'unknown')
            decision_breakdown[decision] = decision_breakdown.get(decision, 0) + 1
        
        return {
            'total_feedback': len(feedback_history),
            'latest_decision': latest_feedback.get('decision') if latest_feedback else None,
            'average_rating': round(sum(ratings) / len(ratings), 2) if ratings else None,
            'latest_feedback_time': latest_feedback.get('timestamp') if latest_feedback else None,
            'decision_breakdown': decision_breakdown  # ✅ Add decision breakdown
        }
        
    except Exception as e:
        print(f"Error getting feedback info for {filename}: {e}")
        return {
            'total_feedback': 0,
            'latest_decision': None,
            'average_rating': None,
            'latest_feedback_time': None,
            'decision_breakdown': {}  # ✅ Add missing field
        }


@bp.route('/download_csv')
def download_csv():
    """
    Download analysis results as CSV with feedback information
    """
    try:
        results = get_ranking_results()
        
        if not results:
            flash('No data to download.', 'error')
            return render_template('index.html')
        
        # Create CSV content with enhanced feedback columns
        csv_content = "Rank,Name,Email,Phone,Years_Experience,Skills,TF-IDF_Score,Semantic_Score,Skill_Score,Combined_Score,Selected,Total_Feedback,Latest_Decision,Average_Rating,Latest_Feedback_Time,Decision_Breakdown\n"
        
        for resume in results:
            selected_status = "Yes" if resume.get('selected') else "No"
            skills_text = resume.get('skills_text', '')
            
            # Get feedback information
            feedback_info = get_cv_feedback_info(resume['filename'])
            
            # Format decision breakdown
            decision_breakdown = feedback_info.get('decision_breakdown', {})
            breakdown_text = ', '.join([f"{k}:{v}" for k, v in decision_breakdown.items()])
            
            # Escape CSV values
            def escape_csv_value(value):
                if value is None:
                    return ""
                str_value = str(value)
                if ',' in str_value or '"' in str_value or '\n' in str_value:
                    return '"' + str_value + '"'
                return str_value
            
            # Create row data
            row_data = [
                resume['rank'],
                escape_csv_value(resume['name']),
                escape_csv_value(resume['email']),
                escape_csv_value(resume['phone']),
                resume['years_exp'],
                escape_csv_value(skills_text),
                f"{resume['tfidf_score']:.2f}",
                f"{resume['semantic_score']:.2f}",
                f"{resume['skill_score']:.2f}",
                f"{resume['combined_score']:.2f}",
                selected_status,
                feedback_info['total_feedback'],
                escape_csv_value(feedback_info['latest_decision']),
                feedback_info['average_rating'] or '',
                escape_csv_value(feedback_info['latest_feedback_time']),
                escape_csv_value(breakdown_text)
            ]
            
            csv_content += ','.join(map(str, row_data)) + '\n'
        
        # Create CSV file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ranked_resumes_{timestamp}.csv"
        
        # Write temporary file
        temp_file = os.path.join("control", filename)
        with open(temp_file, 'w', encoding='utf-8-sig') as f:
            f.write(csv_content)
        
        return send_file(temp_file, as_attachment=True, download_name=filename)
        
    except Exception as e:
        flash(f'Error creating CSV file: {str(e)}', 'error')
        return render_template('index.html')


@bp.route('/download_excel')
def download_excel():
    """
    Download analysis results as Excel with feedback information
    """
    try:
        results = get_ranking_results()
        
        if not results:
            flash('No data to download.', 'error')
            return render_template('index.html')
        
        # Create DataFrame with same structure as PDF
        data = []
        for resume in results:
            # Get feedback information
            feedback_info = get_cv_feedback_info(resume['filename'])
            
            # Format data same as PDF
            selected_text = "✓" if resume.get('selected') else ""
            feedback_count = feedback_info.get('total_feedback', 0)
            latest_decision = feedback_info.get('latest_decision', '') or ''
            avg_rating = feedback_info.get('average_rating', '') or ''
            
            data.append({
                'Rank': resume.get('rank', 0),
                'Name': resume.get('name', ''),
                'Email': resume.get('email', ''),
                'Phone': resume.get('phone', ''),
                'Experience': f"{resume.get('years_exp', 0)}y",
                'TF-IDF': f"{resume.get('tfidf_score', 0):.1f}",
                'Semantic': f"{resume.get('semantic_score', 0):.1f}",
                'Skills': f"{resume.get('skill_score', 0):.1f}",
                'Combined': f"{resume.get('combined_score', 0):.1f}",
                'Selected': selected_text,
                'Feedback': feedback_count,
                'Decision': latest_decision,
                'Rating': avg_rating
            })
        
        df = pd.DataFrame(data)
        
        # ✅ Use BytesIO instead of temporary file (like upload_list)
        from io import BytesIO
        from flask import Response
        from datetime import datetime
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Ranked Resumes', index=False)
            
            # Get worksheet for formatting
            worksheet = writer.sheets['Ranked Resumes']
            
            # Color selected rows yellow
            try:
                from openpyxl.styles import PatternFill
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                for idx, row in df.iterrows():
                    if row['Selected'] == '✓':
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=idx + 2, column=col).fill = yellow_fill
                            
            except Exception as format_error:
                print(f"⚠️ Excel formatting error (non-critical): {format_error}")
        
        output.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ranked_resumes_{timestamp}.xlsx"
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
        
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        flash(f'Error creating Excel file: {str(e)}', 'error')
        return render_template('index.html')


@bp.route('/download_pdf')
def download_pdf():
    """
    Download analysis results as PDF with feedback information (Landscape)
    """
    try:
        results = get_ranking_results()
        
        if not results:
            flash('No data to download.', 'error')
            return render_template('index.html')
        
        # ✅ Use BytesIO instead of temporary file (like upload_list)
        from io import BytesIO
        from flask import Response
        from datetime import datetime
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), 
                              topMargin=0.5*inch, bottomMargin=0.5*inch,
                              leftMargin=0.3*inch, rightMargin=0.3*inch)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            alignment=1,  # Center alignment
            spaceAfter=20
        )
        
        # Title
        title = Paragraph("AI Resume Ranker - Results with Feedback", title_style)
        elements.append(title)
        elements.append(Spacer(1, 10))
        
        # Create table data with full feedback columns (same as Excel)
        table_data = [['Rank', 'Name', 'Email', 'Phone', 'Experience', 'Skills_Text', 'TF-IDF', 'Semantic', 'Skills', 'Combined', 'Selected', 'Total_Feedback', 'Latest_Decision', 'Average_Rating', 'Latest_Feedback_Time', 'Decision_Breakdown']]
        
        for resume in results:
            # Get feedback information
            feedback_info = get_cv_feedback_info(resume['filename'])
            
            # Format decision breakdown as string
            decision_breakdown = feedback_info.get('decision_breakdown', {})
            breakdown_text = ', '.join([f"{k}:{v}" for k, v in decision_breakdown.items()])
            
            # Format data
            selected_text = "✓" if resume.get('selected') else ""
            feedback_count = str(feedback_info.get('total_feedback', 0))
            latest_decision = feedback_info.get('latest_decision', '') or ''
            avg_rating = str(feedback_info.get('average_rating', '')) if feedback_info.get('average_rating') else ''
            latest_feedback_time = feedback_info.get('latest_feedback_time', '') or ''
            skills_text = resume.get('skills_text', '')[:30] + '...' if len(resume.get('skills_text', '')) > 30 else resume.get('skills_text', '')
            
            table_data.append([
                str(resume.get('rank', 0)),
                resume.get('name', '')[:20] + '...' if len(resume.get('name', '')) > 20 else resume.get('name', ''),
                resume.get('email', '')[:18] + '...' if len(resume.get('email', '')) > 18 else resume.get('email', ''),
                resume.get('phone', '')[:12] + '...' if len(resume.get('phone', '')) > 12 else resume.get('phone', ''),
                f"{resume.get('years_exp', 0)}y",
                skills_text,
                f"{resume.get('tfidf_score', 0):.1f}",
                f"{resume.get('semantic_score', 0):.1f}",
                f"{resume.get('skill_score', 0):.1f}",
                f"{resume.get('combined_score', 0):.1f}",
                selected_text,
                feedback_count,
                latest_decision[:12] + '...' if len(latest_decision) > 12 else latest_decision,
                avg_rating,
                latest_feedback_time[:10] + '...' if len(latest_feedback_time) > 10 else latest_feedback_time,
                breakdown_text[:15] + '...' if len(breakdown_text) > 15 else breakdown_text
            ])
        
        # Create table with column widths optimized for landscape (more columns)
        col_widths = [20, 60, 70, 50, 25, 60, 25, 25, 25, 30, 25, 25, 50, 25, 40, 50]
        table = Table(table_data, colWidths=col_widths)
        
        # Table style (optimized for landscape)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), '#4472C4'),
            ('TEXTCOLOR', (0, 0), (-1, 0), '#FFFFFF'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), '#F2F2F2'),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, '#CCCCCC'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ])
        
        # Color selected rows yellow
        for i in range(1, len(table_data)):
            if table_data[i][9] == "✓":  # Selected column
                table_style.add('BACKGROUND', (0, i), (-1, i), '#FFFF00')
        
        table.setStyle(table_style)
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        output.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ranked_resumes_{timestamp}.pdf"
        
        return Response(
            output.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
        
    except Exception as e:
        print(f"❌ Error creating PDF file: {e}")
        flash(f'Error creating PDF file: {str(e)}', 'error')
        return render_template('index.html')


@bp.route('/select_candidate/<int:rank>', methods=['POST'])
def select_candidate(rank: int):
    """
    Select candidate by rank
    """
    try:
        results = get_ranking_results()
        
        print(f" Debug select_candidate:")
        print(f"  - Requested rank: {rank}")
        print(f"  - Results count: {len(results)}")
        print(f"  - Results: {[r.get('filename', 'Unknown') for r in results]}")
        
        if not results:
            print("❌ No ranking results found!")
            return jsonify({'success': False, 'message': 'No candidate data. Please run analysis first.'}), 404
        
        if 1 <= rank <= len(results):
            results[rank - 1]['selected'] = True
            set_ranking_results(results)
            print(f"✅ Selected candidate #{rank}: {results[rank - 1].get('filename', 'Unknown')}")
            return jsonify({'success': True, 'message': f'Successfully selected candidate #{rank}.'})
        else:
            print(f"❌ Invalid rank: {rank} (valid range: 1-{len(results)})")
            return jsonify({'success': False, 'message': f'Invalid rank. Valid range: 1-{len(results)}'}), 404
            
    except Exception as e:
        print(f"❌ Error selecting candidate: {e}")
        return jsonify({'success': False, 'message': 'An error occurred while selecting candidate.'}), 500


@bp.route('/unselect_candidate/<int:rank>', methods=['POST'])
def unselect_candidate(rank: int):
    """
    Unselect candidate by rank
    """
    try:
        results = get_ranking_results()
        
        if not results:
            return jsonify({'success': False, 'message': 'No candidate data.'}), 404
        
        if 1 <= rank <= len(results):
            results[rank - 1]['selected'] = False
            set_ranking_results(results)
            return jsonify({'success': True, 'message': f'Unselected candidate #{rank}.'})
        else:
            return jsonify({'success': False, 'message': 'Invalid rank.'}), 404
            
    except Exception as e:
        print(f"Error unselecting candidate: {e}")
        return jsonify({'success': False, 'message': 'An error occurred while unselecting candidate.'}), 500


@bp.route('/debug_cvs')
@login_required
def debug_cvs():
    """Debug route to check CV data"""
    from services import debug_cv_data
    debug_cv_data()
    return "Debug info printed to console. Check server logs."


@bp.route('/debug_session')
@login_required
def debug_session():
    """Debug session data"""
    try:
        from flask import session
        return jsonify({
            'success': True,
            'session_keys': list(session.keys()),
            'uploaded_cvs_count': len(session.get('uploaded_cvs', [])),
            'has_jd': 'current_jd' in session,
            'jd_info': session.get('current_jd', {}),
            'uploaded_cvs': session.get('uploaded_cvs', [])
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/debug_ranking')
@login_required
def debug_ranking():
    """Debug ranking results"""
    try:
        from services import get_ranking_results, load_uploaded_cvs_from_session
        
        # Check uploaded CVs
        uploaded_cvs = load_uploaded_cvs_from_session()
        
        # Check ranking results
        ranking_results = get_ranking_results()
        
        return jsonify({
            'success': True,
            'uploaded_cvs_count': len(uploaded_cvs),
            'uploaded_cvs': [cv.get('filename', 'Unknown') for cv in uploaded_cvs],
            'ranking_results_count': len(ranking_results),
            'ranking_results': ranking_results,
            'message': f'Found {len(uploaded_cvs)} CVs and {len(ranking_results)} ranking results'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })


@bp.route('/upload_jd', methods=['POST'])
@login_required
def upload_jd():
    """Upload and save JD file to user directory and session"""
    try:
        if 'jd_file' not in request.files:
            return jsonify({'success': False, 'message': 'No JD file selected!'}), 400
        
        jd_file = request.files['jd_file']
        if jd_file.filename == '':
            return jsonify({'success': False, 'message': 'No JD file selected!'}), 400
        
        if not jd_file.filename.lower().endswith(('.pdf', '.docx', '.txt')):
            return jsonify({'success': False, 'message': 'Only PDF, DOCX, TXT files are supported!'}), 400
        
        # ✅ Use user directory system
        from services import get_user_uploads_dir, save_file_to_user_dir
        
        # Save file to user directory
        jd_filename = f"jd_{jd_file.filename}"
        jd_path = save_file_to_user_dir(jd_file, jd_filename)
        
        if not jd_path:
            return jsonify({'success': False, 'message': 'Failed to save JD file!'}), 500
        
        # Extract content
        if jd_file.filename.lower().endswith('.txt'):
            with open(jd_path, 'r', encoding='utf-8') as f:
                content = f.read()
        else:
            content = extract_text(jd_path)
        
        # Save to session
        jd_data = {
            'filename': jd_file.filename,
            'content': content,
            'file_size': os.path.getsize(jd_path),
            'file_path': jd_path  # ✅ Save path for later access
        }
        save_jd_to_session(jd_data)
        
        print(f"✅ JD uploaded successfully: {jd_file.filename}")
        return jsonify({
            'success': True, 
            'message': f'JD uploaded successfully!',
            'content': content,  # ✅ Return content for display
            'filename': jd_file.filename
        })
        
    except Exception as e:
        print(f"❌ Error uploading JD: {e}")
        return jsonify({'success': False, 'message': f'Error uploading JD: {str(e)}'}), 500


@bp.route('/get_jd_info')
@login_required
def get_jd_info():
    """Get JD information from session"""
    try:
        jd_data = get_jd_from_session()
        if jd_data:
            return jsonify({
                'success': True,
                'jd_info': {
                    'filename': jd_data.get('filename', ''),
                    'content_preview': jd_data.get('content', '')[:200] + '...' if len(jd_data.get('content', '')) > 200 else jd_data.get('content', ''),
                    'full_content': jd_data.get('content', ''),
                    'upload_time': jd_data.get('upload_time', ''),
                    'file_size': jd_data.get('file_size', 0),
                    'source': jd_data.get('source', 'file'),
                    'char_count': len(jd_data.get('content', ''))
                }
            })
        else:
            return jsonify({'success': False, 'message': 'No JD has been uploaded'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@bp.route('/get_jd_from_session')
@login_required
def get_jd_from_session():
    """Get JD data from session"""
    try:
        jd_data = get_jd_from_session()
        if jd_data:
            return jsonify({
                'success': True,
                'jd_data': jd_data
            })
        else:
            return jsonify({
                'success': False,
                'message': 'No JD data in session'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        })


@bp.route('/clear_jd', methods=['POST'])
@login_required
def clear_jd():
    """Delete JD from session"""
    try:
        clear_jd_from_session()
        return jsonify({'success': True, 'message': 'JD has been deleted from system'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@bp.route('/upload_cv_files', methods=['POST'])
@login_required
def upload_cv_files():
    """Step 1: Only save CV files to user's uploads directory"""
    try:
        if 'resume_files' not in request.files:
            return jsonify({'success': False, 'message': 'No CV files selected!'}), 400

        resume_files = request.files.getlist('resume_files')
        if not resume_files or all(f.filename == '' for f in resume_files):
            return jsonify({'success': False, 'message': 'No CV files selected!'}), 400

        # Create uploads directory for user
        user_dir = get_user_uploads_dir()

        uploaded_files = []
        total_size = 0

        for resume_file in resume_files:
            if resume_file.filename and resume_file.filename.lower().endswith(('.pdf', '.docx')):
                try:
                    # Save file to user directory
                    resume_path = save_file_to_user_dir(resume_file, resume_file.filename)
                    if resume_path:
                        file_size = os.path.getsize(resume_path)
                        total_size += file_size
                        
                        uploaded_files.append({
                            'filename': resume_file.filename,
                            'size': file_size,
                            'path': resume_path
                        })
                        print(f"✅ Saved file to user directory: {resume_file.filename}")
                    else:
                        print(f"❌ Cannot save file: {resume_file.filename}")
                except Exception as e:
                    print(f"❌ Error saving file {resume_file.filename}: {e}")

        return jsonify({
            'success': True, 
            'message': f'Successfully saved {len(uploaded_files)} CV files!',
            'files': uploaded_files,
            'total_size': total_size
        })

    except Exception as e:
        print(f"❌ Error saving CV files: {e}")
        return jsonify({'success': False, 'message': f'Error saving CV files: {str(e)}'}), 500


@bp.route('/parse_cv_files', methods=['POST'])
@login_required
def parse_cv_files():
    """Step 2: Parse CV files to JSON and save to session"""
    try:
        data = request.get_json()
        filenames = data.get('filenames', [])
        
        if not filenames:
            return jsonify({'success': False, 'message': 'No files to parse!'}), 400

        parsed_count = 0
        parsed_files = []
        errors = []

        for filename in filenames:
            try:
                # Use user-specific file path
                resume_path = get_user_file_path(filename)
                if not os.path.exists(resume_path):
                    errors.append(f"File {filename} does not exist in user directory")
                    continue

                # Extract information from CV
                cv_data = extract_cv_data(filename)
                
                if cv_data:
                    # Save to session
                    save_cv_to_session(cv_data)
                    parsed_count += 1
                    parsed_files.append({
                        'filename': filename,
                        'name': cv_data.get('name', 'Unknown'),
                        'email': cv_data.get('email', 'N/A'),
                        'years_exp': cv_data.get('years_exp', 0)
                    })
                    print(f"✅ Parsed CV: {filename}")
                else:
                    errors.append(f"Cannot extract information from {filename}")

            except Exception as e:
                error_msg = f"Error parsing {filename}: {str(e)}"
                errors.append(error_msg)
                print(f"❌ {error_msg}")

        result_message = f'Successfully parsed {parsed_count}/{len(filenames)} CV files!'
        if errors:
            result_message += f' There are {len(errors)} errors.'

        return jsonify({
            'success': True,
            'message': result_message,
            'parsed_count': parsed_count,
            'total_count': len(filenames),
            'parsed_files': parsed_files,
            'errors': errors
        })

    except Exception as e:
        print(f"❌ Error parsing CV files: {e}")
        return jsonify({'success': False, 'message': f'Error parsing CV files: {str(e)}'}), 500


@bp.route('/get_uploaded_files')
@login_required
def get_uploaded_files():
    """Get list of uploaded files that haven't been parsed yet"""
    try:
        if not os.path.exists("uploads"):
            return jsonify({'success': True, 'files': []})

        # Get list of files in uploads
        all_files = []
        for filename in os.listdir("uploads"):
            file_path = os.path.join("uploads", filename)
            if os.path.isfile(file_path) and filename.lower().endswith(('.pdf', '.docx')):
                file_size = os.path.getsize(file_path)
                all_files.append({
                    'filename': filename,
                    'size': file_size,
                    'upload_time': os.path.getctime(file_path)
                })

        # Sort by upload time
        all_files.sort(key=lambda x: x['upload_time'], reverse=True)

        return jsonify({'success': True, 'files': all_files})

    except Exception as e:
        print(f"Error getting file list: {e}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@bp.route('/save_jd_text', methods=['POST'])
@login_required
def save_jd_text():
    """Save JD text to session"""
    try:
        data = request.get_json()
        jd_content = data.get('content', '').strip()
        
        if not jd_content:
            return jsonify({'success': False, 'message': 'JD content cannot be empty!'}), 400
        
        # Save to session
        jd_data = {
            'filename': 'manual_input.txt',
            'content': jd_content,
            'source': 'text',
            'upload_time': datetime.now().isoformat(),
            'file_size': len(jd_content.encode('utf-8'))
        }
        save_jd_to_session(jd_data)
        
        return jsonify({
            'success': True, 
            'message': 'JD text saved successfully!',
            'jd_info': {
                'filename': 'Manual Input',
                'content_preview': jd_content[:200] + '...' if len(jd_content) > 200 else jd_content,
                'full_content': jd_content,
                'char_count': len(jd_content),
                'source': 'text',
                'upload_time': jd_data['upload_time']
            }
        })
        
    except Exception as e:
        print(f"Error saving JD text: {e}")
        return jsonify({'success': False, 'message': f'Error saving JD text: {str(e)}'}), 500


@bp.route('/get_uploaded_cvs_count')
@login_required
def get_uploaded_cvs_count():
    """Get number of CVs that have been parsed and saved to session"""
    try:
        from services import load_uploaded_cvs_from_session
        cvs = load_uploaded_cvs_from_session()
        return jsonify({
            'success': True,
            'count': len(cvs),
            'cvs': [{'name': cv.get('name', 'Unknown'), 'filename': cv.get('filename', '')} for cv in cvs]
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@bp.route('/sync_cvs_from_uploads')
@login_required
def sync_cvs_from_uploads():
    """Sync all CVs from uploads directory to session"""
    try:
        uploads_dir = "uploads"
        if not os.path.exists(uploads_dir):
            return jsonify({'success': False, 'message': 'Uploads directory does not exist'})
        
        # Get all JSON files in uploads
        json_files = [f for f in os.listdir(uploads_dir) if f.endswith('.json')]
        
        if not json_files:
            return jsonify({'success': False, 'message': 'No CV JSON files in uploads'})
        
        print(f"Found {len(json_files)} CV JSON files to sync")
        
        # Clear old session
        if 'uploaded_cvs' in session:
            session['uploaded_cvs'] = []
        
        synced_count = 0
        for json_file in json_files:
            json_path = os.path.join(uploads_dir, json_file)
            
            try:
                # Read CV data from JSON file
                with open(json_path, 'r', encoding='utf-8') as f:
                    cv_data = json.load(f)
                
                # Create session record
                session_record = {
                    'filename': cv_data.get('filename', ''),
                    'json_file_path': json_path,
                    'upload_time': cv_data.get('upload_time', '')
                }
                
                # Add to session
                if 'uploaded_cvs' not in session:
                    session['uploaded_cvs'] = []
                session['uploaded_cvs'].append(session_record)
                synced_count += 1
                
                print(f"Synced CV: {cv_data.get('name', 'Unknown')} ({cv_data.get('filename', '')}) - {cv_data.get('years_exp', 0)} years")
                
            except Exception as e:
                print(f"Error syncing {json_file}: {e}")
                continue
        
        session.modified = True
        
        return jsonify({
            'success': True, 
            'message': f'Synced {synced_count}/{len(json_files)} CVs to session',
            'synced_count': synced_count,
            'total_files': len(json_files)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error syncing CVs: {str(e)}'}), 500


@bp.route('/get_selected_candidates')
def get_selected_candidates():
    """API to get list of selected candidates"""
    try:
        selected_candidates = session.get('selected_candidates', [])
        return jsonify({
            'success': True,
            'selected_candidates': selected_candidates
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@bp.route('/save_selected_candidates', methods=['POST'])
def save_selected_candidates():
    """API to save list of selected candidates"""
    try:
        data = request.get_json()
        selected_candidates = data.get('selected_candidates', [])
        
        session['selected_candidates'] = selected_candidates
        session.modified = True
        
        return jsonify({
            'success': True,
            'message': f'Saved {len(selected_candidates)} selected candidates'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


